import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
from collections.abc import Callable, Iterable
from contextlib import suppress
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from threading import Event, Lock, Thread
from time import sleep
from uuid import uuid4

from eor_control.domain import MeasurementRecord
from eor_control.project_database import (
    MeasurementQuery,
    ProjectSQLiteWriter,
    create_database_snapshot,
    initialize_project_database,
    list_phases,
    query_events,
    query_measurements,
    write_migration_rows,
)
from eor_control.storage import CsvMeasurementWriter
from eor_control.timezone import as_hungarian_time


def safe_filename(value: str) -> str:
    """Return a Windows-safe, stable file-name component."""
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value.strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned[:80] or "projekt"


@dataclass(frozen=True, slots=True)
class MeasurementTable:
    header: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]

    def column(self, name: str) -> tuple[str, ...]:
        try:
            index = self.header.index(name)
        except ValueError as error:
            raise KeyError(name) from error
        return tuple(row[index] for row in self.rows)


@dataclass(frozen=True, slots=True)
class MeasurementEvent:
    """One durable operator/fault marker shared by every measurement view."""

    event_id: str
    recorded_at_utc: str
    elapsed_seconds: float
    event_type: str
    severity: str
    error_code: str
    description: str
    active_stage: str
    affected_hardware: str
    jacket_pressure_bar: float | None = None
    injection_pressure_bar: float | None = None
    line_pressure_bar: float | None = None
    differential_pressure_bar: float | None = None
    current_flow_ml_per_hour: float | None = None
    target_flow_ml_per_hour: float | None = None
    valve_output_percent: float | None = None
    measurement_state: str = "idle"


def measurement_event_path(source: Path) -> Path:
    return source.with_suffix(".events.jsonl")


def read_measurement_events(paths: Iterable[Path]) -> tuple[MeasurementEvent, ...]:
    events: dict[str, MeasurementEvent] = {}
    for source in dict.fromkeys(paths):
        if source.name == "project.sqlite" and source.is_file():
            for row in query_events(source):
                details = json.loads(str(row["details_json"]))
                try:
                    event = MeasurementEvent(**details)
                except TypeError:
                    event = MeasurementEvent(
                        event_id=str(row["id"]),
                        recorded_at_utc=str(row["recorded_at"]),
                        elapsed_seconds=0.0,
                        event_type=str(row["event_type"]),
                        severity=str(row["severity"]),
                        error_code="",
                        description=str(row["message"]),
                        active_stage="",
                        affected_hardware=str(row["source"]),
                    )
                events[event.event_id] = event
            continue
        path = measurement_event_path(source)
        if not path.is_file():
            continue
        with path.open(encoding="utf-8") as file:
            for line_number, line in enumerate(file, start=1):
                try:
                    payload = json.loads(line)
                    event = MeasurementEvent(**payload)
                except (TypeError, ValueError, json.JSONDecodeError) as error:
                    raise ValueError(
                        f"érvénytelen mérési esemény: {path}:{line_number}: {error}"
                    ) from error
                events[event.event_id] = event
    return tuple(sorted(events.values(), key=lambda event: event.recorded_at_utc))


def measurement_stages(table: MeasurementTable) -> tuple[str, ...]:
    """Return non-empty stage names in first-occurrence order."""

    if "active_stage" not in table.header:
        return ()
    return tuple(dict.fromkeys(value for value in table.column("active_stage") if value))


def filter_measurement_table_by_stage(
    table: MeasurementTable, stage: str | None
) -> MeasurementTable:
    if stage is None or "active_stage" not in table.header:
        return table
    stage_index = table.header.index("active_stage")
    return MeasurementTable(
        table.header,
        tuple(row for row in table.rows if row[stage_index] == stage),
    )


def measurement_stage_segments(
    table: MeasurementTable,
) -> tuple[tuple[str, int, int], ...]:
    """Return contiguous stage spans as (name, start, exclusive end)."""

    if not table.rows or "active_stage" not in table.header:
        return ()
    values = table.column("active_stage")
    segments: list[tuple[str, int, int]] = []
    start = 0
    current = values[0]
    for index, stage in enumerate(values[1:], start=1):
        if stage != current:
            segments.append((current, start, index))
            current = stage
            start = index
    segments.append((current, start, len(values)))
    return tuple(segments)


def read_measurement_table(path: Path) -> MeasurementTable:
    if not path.is_file():
        return MeasurementTable(CsvMeasurementWriter.HEADER, ())
    with path.open(encoding="utf-8", newline="") as file:
        first_line = file.readline()
        file.seek(0)
        delimiter = ";" if ";" in first_line else ","
        rows = list(csv.reader(file, delimiter=delimiter))
    if not rows:
        return MeasurementTable(CsvMeasurementWriter.HEADER, ())
    header = tuple(rows[0])
    legacy_inlet_column = "inlet_pressure_bar"
    if (
        legacy_inlet_column in header
        and tuple(item for item in header if item != legacy_inlet_column)
        == CsvMeasurementWriter.LEGACY_HEADER
    ):
        inlet_index = header.index(legacy_inlet_column)
        rows = [
            [value for index, value in enumerate(row) if index != inlet_index]
            for row in rows
        ]
        header = tuple(rows[0])
    if header in (
        CsvMeasurementWriter.LEGACY_HEADER,
        CsvMeasurementWriter.V2_HEADER,
        CsvMeasurementWriter.V3_HEADER,
        CsvMeasurementWriter.V4_HEADER,
        CsvMeasurementWriter.V5_HEADER,
    ):
        legacy_index = {name: index for index, name in enumerate(header)}
        converted_rows: list[list[str]] = [list(CsvMeasurementWriter.HEADER)]
        for row in rows[1:]:
            converted_rows.append(
                [
                    (
                        ""
                        if name == "jacket_net_volume_ml"
                        and name not in legacy_index
                        else row[legacy_index["injected_volume_ml"]]
                        if name == "injection_net_volume_ml"
                        and name not in legacy_index
                        else row[legacy_index["line_pressure_bar"]]
                        if name == "raw_line_pressure_bar"
                        else row[legacy_index["differential_pressure_bar"]]
                        if name == "raw_differential_pressure_bar"
                        else row[legacy_index["raw_line_voltage"]]
                        if name == "median_line_voltage"
                        and "raw_line_voltage" in legacy_index
                        else row[legacy_index["quality"]]
                        if name == "line_pressure_quality"
                        else row[legacy_index["raw_differential_voltage"]]
                        if name == "median_differential_voltage"
                        and "raw_differential_voltage" in legacy_index
                        else row[legacy_index["quality"]]
                        if name == "differential_pressure_quality"
                        else ""
                        if name
                        in {
                            "raw_line_voltage",
                            "median_line_voltage",
                            "filtered_line_voltage",
                            "line_pressure_quality_reason",
                            "line_pressure_sample_age_seconds",
                            "raw_differential_voltage",
                            "median_differential_voltage",
                            "filtered_differential_voltage",
                            "differential_pressure_quality_reason",
                            "differential_pressure_sample_age_seconds",
                        }
                        else row[legacy_index[name]]
                    )
                    for name in CsvMeasurementWriter.HEADER
                ]
            )
        rows = converted_rows
        header = CsvMeasurementWriter.HEADER
    if header != CsvMeasurementWriter.HEADER:
        raise ValueError("a mérési CSV fejléce nem támogatott")
    width = len(header)
    valid_rows = tuple(tuple(row) for row in rows[1:] if len(row) == width)
    return MeasurementTable(header, valid_rows)


def read_measurement_tables(paths: Iterable[Path]) -> MeasurementTable:
    """Combine phase CSV files in memory without creating a merged data file."""

    rows: list[tuple[str, ...]] = []
    for path in dict.fromkeys(paths):
        if path.name == "project.sqlite" and path.is_file():
            rows.extend(_read_project_measurement_table(path).rows)
        else:
            rows.extend(read_measurement_table(path).rows)

    def recorded_at(row: tuple[str, ...]) -> datetime:
        try:
            return datetime.fromisoformat(row[0].replace("Z", "+00:00")).astimezone(UTC)
        except (IndexError, ValueError):
            return datetime.max.replace(tzinfo=UTC)

    rows.sort(key=recorded_at)
    return MeasurementTable(CsvMeasurementWriter.HEADER, tuple(rows))


def _read_project_measurement_table(
    path: Path,
    *,
    max_points: int | None = 20_000,
    phase_id: int | None = None,
) -> MeasurementTable:
    """Compatibility view for UI models, bounded by database-side downsampling."""
    columns = (
        "recorded_at",
        "project_elapsed_s",
        "jacket_pressure_bar",
        "jacket_flow_ml_min",
        "jacket_remaining_volume_ml",
        "jacket_injected_volume_ml",
        "injection_pressure_bar",
        "injection_flow_ml_min",
        "injection_remaining_volume_ml",
        "injection_injected_volume_ml",
        "raw_line_voltage",
        "median_line_voltage",
        "filtered_line_voltage",
        "raw_line_pressure_bar",
        "line_pressure_bar",
        "line_data_quality",
        "line_quality_reason",
        "line_sample_age_s",
        "raw_differential_voltage",
        "median_differential_voltage",
        "filtered_differential_voltage",
        "raw_differential_pressure_bar",
        "differential_pressure_bar",
        "differential_data_quality",
        "differential_quality_reason",
        "differential_sample_age_s",
        "valve_position_percent",
        "control_state",
        "pressure_data_quality",
        "diagnostic_flags",
    )
    queried = query_measurements(
        path,
        MeasurementQuery(phase_id=phase_id, columns=columns, max_points=max_points),
    )

    def text(value: object) -> str:
        return "" if value is None else str(value)

    converted: list[tuple[str, ...]] = []
    for row in queried.rows:
        values = dict(zip(queried.columns, row, strict=True))
        converted.append(
            (
                text(values["recorded_at"]),
                text(values["project_elapsed_s"]),
                text(values["jacket_pressure_bar"]),
                text(
                    None
                    if values["jacket_flow_ml_min"] is None
                    else float(str(values["jacket_flow_ml_min"])) * 60.0
                ),
                text(values["jacket_remaining_volume_ml"]),
                text(values["jacket_injected_volume_ml"]),
                text(values["injection_pressure_bar"]),
                text(
                    None
                    if values["injection_flow_ml_min"] is None
                    else float(str(values["injection_flow_ml_min"])) * 60.0
                ),
                text(values["injection_remaining_volume_ml"]),
                text(values["injection_injected_volume_ml"]),
                text(values["raw_line_voltage"]),
                text(values["median_line_voltage"]),
                text(values["filtered_line_voltage"]),
                text(values["raw_line_pressure_bar"]),
                text(values["line_pressure_bar"]),
                text(values["line_data_quality"]),
                text(values["line_quality_reason"]),
                text(values["line_sample_age_s"]),
                text(values["raw_differential_voltage"]),
                text(values["median_differential_voltage"]),
                text(values["filtered_differential_voltage"]),
                text(values["raw_differential_pressure_bar"]),
                text(values["differential_pressure_bar"]),
                text(values["differential_data_quality"]),
                text(values["differential_quality_reason"]),
                text(values["differential_sample_age_s"]),
                text(values["valve_position_percent"]),
                text(values["control_state"]),
                text(values["pressure_data_quality"]),
                text(values["diagnostic_flags"]),
            )
        )
    return MeasurementTable(CsvMeasurementWriter.HEADER, tuple(converted))


def export_measurement_csv(
    source: Path,
    destination: Path,
    *,
    decimal_comma: bool = True,
    delimiter: str = ";",
    phase_id: int | None = None,
) -> None:
    if delimiter not in {",", ";", "\t"}:
        raise ValueError("a CSV elválasztó csak vessző, pontosvessző vagy tabulátor lehet")
    table = (
        _read_project_measurement_table(source, max_points=None, phase_id=phase_id)
        if source.name == "project.sqlite"
        else read_measurement_table(source)
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    numeric_columns = set(range(1, table.header.index("active_stage")))
    with destination.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, delimiter=delimiter, lineterminator="\n")
        writer.writerow(table.header)
        for row in table.rows:
            values = list(row)
            for index in numeric_columns:
                if decimal_comma:
                    values[index] = values[index].replace(".", ",")
                else:
                    values[index] = values[index].replace(",", ".")
            writer.writerow(values)


@dataclass(frozen=True, slots=True)
class CsvMigrationReport:
    source_files: int
    source_rows: int
    inserted_rows: int
    duplicate_rows: int
    invalid_rows: int
    warnings: tuple[str, ...]


def migrate_legacy_measurement_csvs(
    database_path: Path,
    *,
    project_id: int,
    sources: Iterable[Path],
    phase_types: dict[str, str] | None = None,
) -> CsvMigrationReport:
    """Import legacy phase CSVs idempotently without deleting their sources."""
    source_files = source_rows = inserted_rows = duplicate_rows = invalid_rows = 0
    warnings: list[str] = []
    for source in dict.fromkeys(sources):
        source_files += 1
        try:
            table = read_measurement_table(source)
        except (OSError, ValueError) as error:
            invalid_rows += 1
            warnings.append(f"{source}: nem olvasható CSV: {error}")
            continue
        source_rows += len(table.rows)
        index = {name: position for position, name in enumerate(table.header)}
        by_stage: dict[str, list[dict[str, object]]] = {}
        for line_number, row in enumerate(table.rows, start=2):
            stage = row[index["active_stage"]].strip()
            if not stage:
                invalid_rows += 1
                warnings.append(
                    f"{source}:{line_number}: a szakaszkapcsolat nem állapítható meg"
                )
                continue
            try:
                recorded_at = datetime.fromisoformat(
                    row[index["recorded_at_utc"]].replace("Z", "+00:00")
                ).astimezone(UTC).isoformat()
            except ValueError:
                invalid_rows += 1
                warnings.append(f"{source}:{line_number}: érvénytelen időbélyeg")
                continue
            payload: dict[str, object] = {"recorded_at": recorded_at}
            row_invalid = False
            for name in table.header[1:]:
                value = row[index[name]].strip()
                if name in {
                    "active_stage",
                    "quality",
                    "safety_reasons",
                    "line_pressure_quality",
                    "line_pressure_quality_reason",
                    "differential_pressure_quality",
                    "differential_pressure_quality_reason",
                }:
                    payload[name] = value
                elif value:
                    try:
                        payload[name] = float(value.replace(",", "."))
                    except ValueError:
                        row_invalid = True
                        warnings.append(
                            f"{source}:{line_number}: érvénytelen szám a(z) {name} mezőben"
                        )
                        break
                else:
                    payload[name] = None
            if row_invalid:
                invalid_rows += 1
                continue
            by_stage.setdefault(stage, []).append(payload)
        if len(by_stage) > 1:
            warnings.append(
                f"{source}: több szakasz található egy fájlban; külön példányokba kerültek"
            )
        for stage, rows in by_stage.items():
            phase_type = (phase_types or {}).get(stage, "unknown")
            if phase_type == "unknown":
                warnings.append(
                    f"{source}: a(z) {stage!r} szakasztípus nem ismert; 'unknown' jelölést kapott"
                )
            inserted, duplicates = write_migration_rows(
                database_path,
                project_id=project_id,
                phase_name=stage,
                phase_type=phase_type,
                source_key=f"csv:{source.resolve()}:{stage}",
                rows=rows,
            )
            inserted_rows += inserted
            duplicate_rows += duplicates
    if source_rows != inserted_rows + duplicate_rows + invalid_rows:
        warnings.append(
            "a forrás- és célsorszám eltér; ellenőrizd a migrációs figyelmeztetéseket"
        )
    return CsvMigrationReport(
        source_files,
        source_rows,
        inserted_rows,
        duplicate_rows,
        invalid_rows,
        tuple(warnings),
    )


def project_excel_path(source: Path, stage_name: str) -> Path:
    """Return the project's rebuilt workbook path."""
    if source.name == "project.sqlite":
        project_name = safe_filename(source.parent.name.split("_", 2)[-1])
        return source.with_name(f"{project_name}.xlsx")
    stage = safe_filename(stage_name)
    live_suffix = f"_{stage}_live_raw.csv"
    simulation_suffix = f"_{stage}_simulation_live_raw.csv"
    if source.name.endswith(simulation_suffix):
        project_name = source.name[: -len(simulation_suffix)]
        workbook_suffix = "_simulation"
    elif source.name.endswith(live_suffix):
        project_name = source.name[: -len(live_suffix)]
        workbook_suffix = ""
    else:
        raise ValueError(
            "a mérési szakasz fájlneve nem illeszkedik a projekt-exporthoz"
        )
    if not project_name:
        raise ValueError("a projekt Excel-fájlneve nem lehet üres")
    return source.with_name(f"{project_name}{workbook_suffix}.xlsx")


def _excel_sheet_title(stage_name: str) -> str:
    requested = stage_name.strip() or "Mérés"
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", requested).strip("'") or "Mérés"
    if cleaned == requested and len(cleaned) <= 31:
        return cleaned
    digest = hashlib.sha256(requested.encode("utf-8")).hexdigest()[:6]
    return f"{cleaned[:24]}_{digest}"


def _export_measurement_excel_legacy(
    source: Path,
    destination: Path,
    *,
    stage_name: str,
) -> None:
    """Create or update one stage worksheet in a project workbook."""
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.chart import LineChart, Reference  # type: ignore[import-untyped]
        from openpyxl.chart.series_factory import SeriesFactory  # type: ignore[import-untyped]
        from openpyxl.reader.excel import load_workbook  # type: ignore[import-untyped]
    except ImportError as error:
        raise RuntimeError(
            "Az Excel-exporthoz telepítsd az export függőséget: pip install -e \".[export]\""
        ) from error

    table = read_measurement_table(source)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.is_file():
        workbook = load_workbook(destination)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    sheet_title = _excel_sheet_title(stage_name)
    sheet_index = len(workbook.worksheets)
    if sheet_title in workbook.sheetnames:
        existing = workbook[sheet_title]
        sheet_index = workbook.index(existing)
        workbook.remove(existing)
    sheet = workbook.create_sheet(sheet_title, sheet_index)
    sheet.freeze_panes = "A2"
    sheet.append(list(table.header))
    numeric_columns = set(range(1, table.header.index("active_stage")))
    for source_row in table.rows:
        row: list[str | float] = list(source_row)
        for index in numeric_columns:
            with suppress(ValueError):
                row[index] = float(source_row[index].replace(",", "."))
        sheet.append(row)
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(42, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[letter].width = width

    if table.rows:
        chart = LineChart()
        chart.title = f"{stage_name} — nyomás- és szelepdiagram"
        chart.y_axis.title = "bar / %"
        chart.x_axis.title = "Minta"
        categories = Reference(sheet, min_col=1, min_row=2, max_row=len(table.rows) + 1)
        for column_name in (
            "jacket_pressure_bar",
            "injection_pressure_bar",
            "line_pressure_bar",
            "differential_pressure_bar",
            "valve_percent",
        ):
            column_index = table.header.index(column_name) + 1
            data = Reference(
                sheet,
                min_col=column_index,
                max_col=column_index,
                min_row=1,
                max_row=len(table.rows) + 1,
            )
            chart.add_data(data, titles_from_data=True, from_rows=False)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 24
        sheet.add_chart(chart, "T2")

    events = read_measurement_events((source,))
    event_title = _excel_sheet_title(f"{stage_name} események")
    if event_title in workbook.sheetnames:
        workbook.remove(workbook[event_title])
    if events:
        event_sheet = workbook.create_sheet(event_title, sheet_index + 1)
        event_header = tuple(MeasurementEvent.__dataclass_fields__)
        event_sheet.append(event_header)
        for event in events:
            event_sheet.append([getattr(event, name) for name in event_header])
        event_sheet.freeze_panes = "A2"
        event_sheet.auto_filter.ref = event_sheet.dimensions
        for column in event_sheet.columns:
            event_sheet.column_dimensions[column[0].column_letter].width = min(
                42, max(len(str(cell.value or "")) for cell in column) + 2
            )
    if table.rows and events:
        chart = sheet._charts[0]
        elapsed_column = event_header.index("elapsed_seconds") + 1
        marker_column = event_header.index("injection_pressure_bar") + 1
        values = Reference(
            event_sheet,
            min_col=marker_column,
            min_row=2,
            max_row=len(events) + 1,
        )
        categories = Reference(
            event_sheet,
            min_col=elapsed_column,
            min_row=2,
            max_row=len(events) + 1,
        )
        marker_series = SeriesFactory(values, title="Események", xvalues=categories)
        marker_series.graphicalProperties.noFill = True
        marker_series.marker.symbol = "diamond"
        marker_series.marker.size = 9
        chart.series.append(marker_series)

    temporary = destination.with_name(
        f".{destination.stem}.{uuid4().hex}.tmp{destination.suffix}"
    )
    workbook.save(temporary)
    os.replace(temporary, destination)


EXCEL_MEASUREMENT_HEADERS = (
    "Rögzített idő [UTC]",
    "Indítás óta eltelt idő [h]:mm:ss",
    "Köpenynyomás [bar]",
    "Köpeny térfogatáram [mL/min]",
    "Köpeny visszamaradt térfogat [mL]",
    "Köpeny besajtolt térfogat [mL]",
    "Besajtolónyomás [bar]",
    "Besajtoló térfogatáram [mL/min]",
    "Besajtolásból megmaradt térfogat [mL]",
    "Besajtoló besajtolt térfogat [mL]",
    "Vonali nyomás [bar]",
    "Differenciálnyomás [bar]",
    "Szelepállás [%]",
)


def export_measurement_excel(
    source: Path,
    destination: Path,
    *,
    stage_name: str = "",
) -> None:
    """Rebuild the complete project workbook exclusively from SQLite."""
    del stage_name  # Kept in the signature while existing UI calls transition.
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.reader.excel import load_workbook
        from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
    except ImportError as error:
        raise RuntimeError(
            "Az Excel-exporthoz telepítsd az export függőséget: pip install -e \".[export]\""
        ) from error
    if source.name != "project.sqlite":
        raise ValueError("az Excel-export kizárólagos forrása a project.sqlite lehet")
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    columns = (
        "recorded_at",
        "project_elapsed_s",
        "jacket_pressure_bar",
        "jacket_flow_ml_min",
        "jacket_remaining_volume_ml",
        "jacket_injected_volume_ml",
        "injection_pressure_bar",
        "injection_flow_ml_min",
        "injection_remaining_volume_ml",
        "injection_injected_volume_ml",
        "line_pressure_bar",
        "differential_pressure_bar",
        "valve_position_percent",
    )
    phases = list_phases(source)
    used_titles: set[str] = set()
    for phase in phases:
        base_title = _excel_sheet_title(f"{phase.sequence:02d} {phase.name}")
        sheet_title = base_title
        duplicate = 2
        while sheet_title.casefold() in used_titles:
            suffix = f" ({duplicate})"
            sheet_title = f"{base_title[: 31 - len(suffix)]}{suffix}"
            duplicate += 1
        used_titles.add(sheet_title.casefold())
        sheet = workbook.create_sheet(sheet_title)
        sheet.freeze_panes = "A2"
        sheet.append(EXCEL_MEASUREMENT_HEADERS)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="548235")
            cell.font = Font(color="FFFFFF", bold=True)
        rows = query_measurements(
            source,
            MeasurementQuery(phase_id=phase.id, columns=columns),
        )
        for values in rows.rows:
            timestamp = datetime.fromisoformat(str(values[0]).replace("Z", "+00:00"))
            sheet.append(
                (
                    timestamp.replace(tzinfo=None),
                    float(str(values[1])) / 86400.0,
                    *values[2:],
                )
            )
        sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 25
        for column in range(3, 14):
            sheet.column_dimensions[sheet.cell(1, column).column_letter].width = 24
        for cell in sheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
        for cell in sheet["B"][1:]:
            cell.number_format = "[h]:mm:ss"
        if rows.rows:
            chart = LineChart()
            chart.title = f"{phase.name} — nyomás- és szelepdiagram"
            chart.y_axis.title = "bar / %"
            chart.x_axis.title = "Indítás óta eltelt idő"
            chart.set_categories(
                Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
            )
            for column in (3, 7, 11, 12, 13):
                chart.add_data(
                    Reference(
                        sheet,
                        min_col=column,
                        max_col=column,
                        min_row=1,
                        max_row=sheet.max_row,
                    ),
                    titles_from_data=True,
                )
            chart.height = 10
            chart.width = 24
            sheet.add_chart(chart, "O2")
    if not phases:
        sheet = workbook.create_sheet("Nincs mérési adat")
        sheet.append(EXCEL_MEASUREMENT_HEADERS)
    temporary = destination.with_name(
        f".{destination.stem}.{uuid4().hex}.tmp{destination.suffix}"
    )
    try:
        workbook.save(temporary)
        verified = load_workbook(temporary, read_only=True, data_only=True)
        try:
            if len(verified.sheetnames) != max(1, len(phases)):
                raise RuntimeError("az Excel-export visszaellenőrzése sikertelen")
            for sheet in verified.worksheets:
                if tuple(cell.value for cell in sheet[1]) != EXCEL_MEASUREMENT_HEADERS:
                    raise RuntimeError("az Excel-export fejlécének ellenőrzése sikertelen")
        finally:
            verified.close()
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)


@dataclass(frozen=True, slots=True)
class NasQueueItem:
    relative_path: str
    source_path: Path
    revision: int
    attempts: int
    last_error: str


@dataclass(frozen=True, slots=True)
class NasConnectionTestResult:
    target_root: Path
    writable: bool
    free_bytes: int
    visible_entries: tuple[str, ...]


def test_nas_connection(target_root: Path) -> NasConnectionTestResult:
    """Verify an existing NAS folder using the current Windows credentials.

    The probe performs a real create/read/delete round trip.  It never creates
    the target directory and never stores authentication material.
    """
    target = target_root.expanduser()
    if not target.exists():
        raise ConnectionError(f"a NAS célmappa nem érhető el: {target}")
    if not target.is_dir():
        raise NotADirectoryError(f"a NAS célútvonal nem mappa: {target}")
    try:
        visible_entries = tuple(
            sorted((item.name for item in target.iterdir()), key=str.casefold)[:100]
        )
        free_bytes = shutil.disk_usage(target).free
    except OSError as error:
        raise ConnectionError(f"a NAS mappa nem olvasható: {error}") from error

    probe = target / f".eor-write-test-{uuid4().hex}.tmp"
    try:
        with probe.open("x+b") as file:
            payload = b"EOR NAS write test\n"
            file.write(payload)
            file.flush()
            os.fsync(file.fileno())
            file.seek(0)
            if file.read() != payload:
                raise OSError("a NAS próba visszaolvasott tartalma eltér")
    except OSError as error:
        raise PermissionError(f"a NAS célmappa nem írható: {error}") from error
    finally:
        try:
            probe.unlink(missing_ok=True)
        except OSError as error:
            raise PermissionError(
                f"a NAS próbfájl nem távolítható el: {probe}: {error}"
            ) from error
    return NasConnectionTestResult(target, True, free_bytes, visible_entries)


class NasSyncQueue:
    """Persistent SQLite queue; survives application and network failures."""

    def __init__(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        self._connection = sqlite3.connect(path, check_same_thread=False)
        self._lock = Lock()
        with self._lock, self._connection:
            self._connection.execute(
                """
                CREATE TABLE IF NOT EXISTS nas_sync_queue (
                    relative_path TEXT PRIMARY KEY,
                    source_path TEXT NOT NULL,
                    queued_at_utc TEXT NOT NULL,
                    revision INTEGER NOT NULL DEFAULT 1,
                    attempts INTEGER NOT NULL DEFAULT 0,
                    last_error TEXT NOT NULL DEFAULT ''
                )
                """
            )
            columns = {
                str(row[1])
                for row in self._connection.execute("PRAGMA table_info(nas_sync_queue)")
            }
            if "revision" not in columns:
                self._connection.execute(
                    "ALTER TABLE nas_sync_queue ADD COLUMN revision INTEGER NOT NULL DEFAULT 1"
                )

    def enqueue(self, source_path: Path, relative_path: Path) -> None:
        relative = _validated_relative_path(relative_path)
        with self._lock, self._connection:
            self._connection.execute(
                """
                INSERT INTO nas_sync_queue (
                    relative_path, source_path, queued_at_utc, revision, attempts, last_error
                ) VALUES (?, ?, ?, 1, 0, '')
                ON CONFLICT(relative_path) DO UPDATE SET
                    source_path=excluded.source_path,
                    queued_at_utc=excluded.queued_at_utc,
                    revision=nas_sync_queue.revision + 1
                """,
                (relative.as_posix(), str(source_path.resolve()), datetime.now(UTC).isoformat()),
            )

    def pending(self) -> tuple[NasQueueItem, ...]:
        with self._lock:
            rows = self._connection.execute(
                """
                SELECT relative_path, source_path, revision, attempts, last_error
                FROM nas_sync_queue ORDER BY queued_at_utc
                """
            ).fetchall()
        return tuple(
            NasQueueItem(
                str(row[0]), Path(str(row[1])), int(row[2]), int(row[3]), str(row[4])
            )
            for row in rows
        )

    def complete(self, relative_path: str, revision: int) -> None:
        with self._lock, self._connection:
            self._connection.execute(
                "DELETE FROM nas_sync_queue WHERE relative_path = ? AND revision = ?",
                (relative_path, revision),
            )

    def fail(self, relative_path: str, message: str) -> None:
        with self._lock, self._connection:
            self._connection.execute(
                """
                UPDATE nas_sync_queue
                SET attempts = attempts + 1, last_error = ?
                WHERE relative_path = ?
                """,
                (message[:1000], relative_path),
            )

    def close(self) -> None:
        with self._lock:
            self._connection.close()


def _validated_relative_path(path: Path) -> Path:
    if path.is_absolute() or ".." in path.parts or not path.parts:
        raise ValueError("a NAS célútvonalnak biztonságos relatív útvonalnak kell lennie")
    return path


class BackgroundNasSynchronizer:
    def __init__(
        self,
        queue: NasSyncQueue,
        *,
        retry_interval_seconds: float = 30.0,
    ) -> None:
        if retry_interval_seconds <= 0.0:
            raise ValueError("a NAS újrapróbálkozási időnek pozitívnak kell lennie")
        self._queue = queue
        self._retry_interval = retry_interval_seconds
        self._target_root: Path | None = None
        self._enabled = False
        self._stop = Event()
        self._wake = Event()
        self._thread: Thread | None = None
        self._lock = Lock()
        self._sync_lock = Lock()

    @property
    def enabled(self) -> bool:
        with self._lock:
            return self._enabled

    @property
    def target_root(self) -> Path | None:
        with self._lock:
            return self._target_root

    @property
    def pending_count(self) -> int:
        return len(self._queue.pending())

    @property
    def pending_errors(self) -> tuple[str, ...]:
        return tuple(
            item.last_error for item in self._queue.pending() if item.last_error
        )

    def configure(self, *, enabled: bool, target_root: Path | None) -> None:
        if enabled and target_root is None:
            raise ValueError("engedélyezett NAS-mentéshez célmappa szükséges")
        with self._lock:
            self._enabled = enabled
            self._target_root = target_root
        if enabled:
            self.start()
            self._wake.set()

    def start(self) -> None:
        if self._thread is not None and self._thread.is_alive():
            return
        self._stop.clear()
        self._thread = Thread(target=self._run, name="eor-nas-sync", daemon=True)
        self._thread.start()

    def enqueue(self, source_path: Path, relative_path: Path) -> None:
        if not self.enabled:
            return
        self._queue.enqueue(source_path, relative_path)
        self.start()
        self._wake.set()

    def sync_pending_once(self) -> int:
        with self._sync_lock:
            with self._lock:
                enabled = self._enabled
                target_root = self._target_root
            if not enabled or target_root is None:
                return 0
            completed = 0
            for item in self._queue.pending():
                try:
                    if not item.source_path.is_file():
                        raise FileNotFoundError(
                            f"forrásfájl nem található: {item.source_path}"
                        )
                    relative = _validated_relative_path(Path(item.relative_path))
                    destination = target_root / relative
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    temporary = destination.with_name(f".{destination.name}.eor-sync.tmp")
                    shutil.copy2(item.source_path, temporary)
                    os.replace(temporary, destination)
                    self._queue.complete(item.relative_path, item.revision)
                    completed += 1
                except OSError as error:
                    self._queue.fail(item.relative_path, str(error))
            return completed

    def _run(self) -> None:
        while not self._stop.is_set():
            self.sync_pending_once()
            self._wake.wait(self._retry_interval)
            self._wake.clear()

    def close(self) -> None:
        self._stop.set()
        self._wake.set()
        thread = self._thread
        if thread is not None:
            thread.join(timeout=2.0)
        self._thread = None
        self._queue.close()


class _LegacyProjectMeasurementWriter:
    """Routes raw records to one crash-safe file per project measurement phase."""

    def __init__(
        self,
        data_root: Path,
        nas_sync: BackgroundNasSynchronizer | None = None,
        *,
        enabled: bool = True,
        measurement_kind: str = "live",
        phase_completed: Callable[[Path, str], None] | None = None,
    ) -> None:
        if measurement_kind not in {"live", "simulation"}:
            raise ValueError("unsupported measurement kind")
        self._data_root = data_root
        self._nas_sync = nas_sync
        self._enabled = enabled
        self._measurement_kind = measurement_kind
        self._phase_completed = phase_completed
        self._writer: CsvMeasurementWriter | None = None
        self._path: Path | None = None
        self._relative_path: Path | None = None
        self._relative_folder: Path | None = None
        self._project_file_prefix: str | None = None
        self._stage_name: str | None = None
        self._phase_has_records = False
        self._lock = Lock()

    @property
    def current_path(self) -> Path | None:
        with self._lock:
            return self._path if self._enabled else None

    @property
    def persistence_enabled(self) -> bool:
        return self._enabled

    def set_phase_completed_callback(
        self, callback: Callable[[Path, str], None] | None
    ) -> None:
        """Attach the UI export callback after the dashboard has been constructed."""
        with self._lock:
            self._phase_completed = callback

    @property
    def phase_paths(self) -> tuple[Path, ...]:
        with self._lock:
            relative_folder = self._relative_folder
        if relative_folder is None:
            return ()
        folder = self._data_root / relative_folder
        return tuple(sorted(folder.glob("*_live_raw.csv")))

    def select_project(
        self, project_id: int, project_name: str, *, stage_name: str = "Mérés"
    ) -> Path:
        return self.select_project_with_metadata(
            project_id, project_name, stage_name=stage_name
        )

    def select_project_with_metadata(
        self,
        project_id: int,
        project_name: str,
        *,
        created_at: datetime | None = None,
        notes: str = "",
        configuration: dict[str, object] | None = None,
        calibration_snapshot: dict[str, object] | None = None,
        stages: list[dict[str, object]] | None = None,
        stage_name: str = "Mérés",
    ) -> Path:
        if project_id <= 0:
            raise ValueError("érvénytelen projektazonosító")
        timestamp = created_at or datetime.now(UTC)
        local_timestamp = as_hungarian_time(timestamp)
        folder = (
            f"{local_timestamp:%Y-%m-%d}_{project_id:06d}_{safe_filename(project_name)}"
        )
        relative_folder = Path("projects") / str(local_timestamp.year) / folder
        project_file_prefix = safe_filename(project_name)
        relative_path = relative_folder / self._raw_filename(
            project_file_prefix, stage_name
        )
        path = self._data_root / relative_path
        completed_phase: tuple[Path, str] | None = None
        with self._lock:
            if path == self._path and (self._writer is not None or not self._enabled):
                return path
            if self._writer is not None:
                self._writer.close()
                if (
                    self._phase_has_records
                    and self._path is not None
                    and self._stage_name is not None
                ):
                    completed_phase = (self._path, self._stage_name)
            self._writer = CsvMeasurementWriter(path) if self._enabled else None
            self._path = path
            self._relative_path = relative_path
            self._relative_folder = relative_folder
            self._project_file_prefix = project_file_prefix
            self._stage_name = stage_name
            self._phase_has_records = False
        if completed_phase is not None and self._phase_completed is not None:
            self._phase_completed(*completed_phase)
        if not self._enabled:
            return path
        self._write_project_snapshots(
            relative_folder,
            project_id=project_id,
            project_name=project_name,
            created_at=timestamp,
            notes=notes,
            configuration=configuration or {},
            calibration_snapshot=calibration_snapshot or {},
            stages=stages or [],
        )
        return path

    def _write_project_snapshots(
        self,
        relative_folder: Path,
        *,
        project_id: int,
        project_name: str,
        created_at: datetime,
        notes: str,
        configuration: dict[str, object],
        calibration_snapshot: dict[str, object],
        stages: list[dict[str, object]],
    ) -> None:
        marker = "" if self._measurement_kind == "live" else "_simulation"
        documents = {
            f"project{marker}.json": {
                "id": project_id,
                "name": project_name,
                "measurement_kind": self._measurement_kind,
                "created_at_utc": created_at.astimezone(UTC).isoformat(),
                "notes": notes,
                "stages": stages,
            },
            f"config_snapshot{marker}.json": configuration,
            f"calibration_snapshot{marker}.json": calibration_snapshot,
        }
        folder = self._data_root / relative_folder
        folder.mkdir(parents=True, exist_ok=True)
        for filename, payload in documents.items():
            destination = folder / filename
            temporary = destination.with_suffix(f"{destination.suffix}.tmp")
            temporary.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
                encoding="utf-8",
            )
            os.replace(temporary, destination)
            if self._nas_sync is not None:
                self._nas_sync.enqueue(
                    destination, Path(*relative_folder.parts[1:]) / filename
                )

    def write(self, record: MeasurementRecord) -> None:
        if not self._enabled:
            return
        completed_phase: tuple[Path, str] | None = None
        with self._lock:
            if self._writer is None or self._path is None or self._relative_path is None:
                raise RuntimeError("a mérés előtt projektet kell kiválasztani")
            if record.active_stage != self._stage_name:
                completed_phase = self._open_phase_locked(record.active_stage)
            assert self._writer is not None
            assert self._path is not None
            assert self._relative_path is not None
            self._writer.write(record)
            self._phase_has_records = True
            path = self._path
            relative_path = self._relative_path
        if completed_phase is not None and self._phase_completed is not None:
            self._phase_completed(*completed_phase)
        if self._nas_sync is not None:
            self._nas_sync.enqueue(path, Path(*relative_path.parts[1:]))

    def write_event(self, event: MeasurementEvent) -> Path:
        """Append a single durable event without duplicating it per diagram."""
        if not self._enabled:
            raise RuntimeError("a mérési eseménymentés nincs engedélyezve")
        with self._lock:
            if self._path is None or self._relative_path is None:
                raise RuntimeError("a mérési esemény előtt projektet kell kiválasztani")
            path = measurement_event_path(self._path)
            relative_path = measurement_event_path(self._relative_path)
            existing_ids = {
                item.event_id for item in read_measurement_events((self._path,))
            }
            if event.event_id not in existing_ids:
                path.parent.mkdir(parents=True, exist_ok=True)
                with path.open("a", encoding="utf-8", newline="\n") as file:
                    file.write(json.dumps(asdict(event), ensure_ascii=False) + "\n")
                    file.flush()
                    os.fsync(file.fileno())
        if self._nas_sync is not None:
            self._nas_sync.enqueue(path, Path(*relative_path.parts[1:]))
        return path

    def _open_phase_locked(self, stage_name: str) -> tuple[Path, str] | None:
        if self._relative_folder is None or self._project_file_prefix is None:
            raise RuntimeError("a mérés előtt projektet kell kiválasztani")
        completed_phase = (
            (self._path, self._stage_name)
            if (
                self._phase_has_records
                and self._path is not None
                and self._stage_name is not None
            )
            else None
        )
        relative_path = self._relative_folder / self._raw_filename(
            self._project_file_prefix, stage_name
        )
        path = self._data_root / relative_path
        if self._writer is not None:
            self._writer.close()
        self._writer = CsvMeasurementWriter(path)
        self._path = path
        self._relative_path = relative_path
        self._stage_name = stage_name
        self._phase_has_records = False
        return completed_phase

    def _raw_filename(self, project_file_prefix: str, stage_name: str) -> str:
        marker = "_simulation" if self._measurement_kind == "simulation" else ""
        return (
            f"{project_file_prefix}_{safe_filename(stage_name)}"
            f"{marker}_live_raw.csv"
        )

    def complete_current_phase(self) -> Path | None:
        completed_phase: tuple[Path, str] | None = None
        with self._lock:
            if self._writer is not None:
                self._writer.close()
                self._writer = None
            if (
                self._phase_has_records
                and self._path is not None
                and self._stage_name is not None
            ):
                completed_phase = (self._path, self._stage_name)
            self._phase_has_records = False
        if completed_phase is not None and self._phase_completed is not None:
            self._phase_completed(*completed_phase)
        return completed_phase[0] if completed_phase is not None else None

    def close(self) -> None:
        with self._lock:
            if self._writer is not None:
                self._writer.close()
            self._writer = None
            self._phase_has_records = False


class ProjectMeasurementWriter:
    """Selects a project database and exposes a non-blocking measurement sink."""

    def __init__(
        self,
        data_root: Path,
        nas_sync: BackgroundNasSynchronizer | None = None,
        *,
        enabled: bool = True,
        measurement_kind: str = "live",
        phase_completed: Callable[[Path, str], None] | None = None,
    ) -> None:
        if measurement_kind not in {"live", "simulation"}:
            raise ValueError("unsupported measurement kind")
        self._data_root = data_root
        self._nas_sync = nas_sync
        self._enabled = enabled
        self._measurement_kind = measurement_kind
        self._phase_completed = phase_completed
        self._writer: ProjectSQLiteWriter | None = None
        self._path: Path | None = None
        self._relative_folder: Path | None = None
        self._stage_name: str | None = None
        self._phase_has_records = False
        self._lock = Lock()

    @property
    def current_path(self) -> Path | None:
        with self._lock:
            return self._path if self._enabled else None

    @property
    def persistence_enabled(self) -> bool:
        return self._enabled

    @property
    def phase_paths(self) -> tuple[Path, ...]:
        path = self.current_path
        return () if path is None else (path,)

    @property
    def queue_metrics(self) -> object | None:
        with self._lock:
            return None if self._writer is None else self._writer.metrics

    def set_phase_completed_callback(
        self, callback: Callable[[Path, str], None] | None
    ) -> None:
        with self._lock:
            self._phase_completed = callback

    def select_project(
        self, project_id: int, project_name: str, *, stage_name: str = "Mérés"
    ) -> Path:
        return self.select_project_with_metadata(
            project_id,
            project_name,
            stage_name=stage_name,
        )

    def select_project_with_metadata(
        self,
        project_id: int,
        project_name: str,
        *,
        created_at: datetime | None = None,
        notes: str = "",
        configuration: dict[str, object] | None = None,
        calibration_snapshot: dict[str, object] | None = None,
        stages: list[dict[str, object]] | None = None,
        stage_name: str = "Mérés",
    ) -> Path:
        if project_id <= 0:
            raise ValueError("érvénytelen projektazonosító")
        timestamp = created_at or datetime.now(UTC)
        local_timestamp = as_hungarian_time(timestamp)
        folder = (
            f"{local_timestamp:%Y-%m-%d}_{project_id:06d}_{safe_filename(project_name)}"
        )
        if self._measurement_kind == "simulation":
            folder += "_simulation"
        relative_folder = Path("projects") / str(local_timestamp.year) / folder
        path = self._data_root / relative_folder / "project.sqlite"
        if not self._enabled:
            return path
        (path.parent / "exports").mkdir(parents=True, exist_ok=True)
        (path.parent / "logs").mkdir(exist_ok=True)
        previous_writer: ProjectSQLiteWriter | None = None
        with self._lock:
            if path == self._path and self._writer is not None:
                self._stage_name = stage_name
                return path
            previous_writer = self._writer
            self._writer = None
        if previous_writer is not None:
            previous_writer.close()
        settings = dict(configuration or {})
        settings["calibration_snapshot"] = calibration_snapshot or {}
        settings["configured_stages"] = stages or []
        settings["measurement_kind"] = self._measurement_kind
        initialize_project_database(
            path,
            project_id=project_id,
            project_name=project_name,
            created_at=timestamp,
            notes=notes,
            settings=settings,
        )
        legacy_sources = tuple(sorted(path.parent.glob("*_live_raw.csv")))
        if legacy_sources:
            phase_types = {
                str(item.get("name", "")): str(item.get("type", "unknown"))
                for item in (stages or [])
                if item.get("name")
            }
            migration_report = migrate_legacy_measurement_csvs(
                path,
                project_id=project_id,
                sources=legacy_sources,
                phase_types=phase_types,
            )
            report_path = path.parent / "logs" / "csv-migration-report.json"
            temporary_report = report_path.with_suffix(".json.tmp")
            temporary_report.write_text(
                json.dumps(
                    asdict(migration_report),
                    ensure_ascii=False,
                    indent=2,
                    sort_keys=True,
                ),
                encoding="utf-8",
            )
            os.replace(temporary_report, report_path)
        writer = ProjectSQLiteWriter(path)
        with self._lock:
            self._writer = writer
            self._path = path
            self._relative_folder = relative_folder
            self._stage_name = stage_name
            self._phase_has_records = False
        return path

    def write(self, record: MeasurementRecord) -> None:
        if not self._enabled:
            return
        with self._lock:
            writer = self._writer
            if writer is None:
                raise RuntimeError("a mérés előtt projektet kell kiválasztani")
            self._stage_name = record.active_stage
            self._phase_has_records = True
        writer.write(record)

    def write_event(self, event: MeasurementEvent) -> Path:
        if not self._enabled:
            raise RuntimeError("a mérési eseménymentés nincs engedélyezve")
        with self._lock:
            writer = self._writer
            path = self._path
        if writer is None or path is None:
            raise RuntimeError("a mérési esemény előtt projektet kell kiválasztani")
        payload = asdict(event)
        writer.write_event(
            event_id=event.event_id,
            recorded_at=event.recorded_at_utc,
            severity=event.severity,
            event_type=event.event_type,
            source=event.affected_hardware or "measurement",
            message=event.description,
            details=payload,
        )
        return path

    def complete_current_phase(self) -> Path | None:
        with self._lock:
            writer = self._writer
            path = self._path
            stage_name = self._stage_name
            has_records = self._phase_has_records
            self._phase_has_records = False
        if writer is None or path is None or not has_records:
            return None
        writer.complete_phase()
        writer.flush()
        self._enqueue_nas_snapshot(path)
        if self._phase_completed is not None and stage_name is not None:
            self._phase_completed(path, stage_name)
        return path

    def _enqueue_nas_snapshot(self, path: Path) -> None:
        if self._nas_sync is None or not self._nas_sync.enabled:
            return
        snapshot_folder = path.parent / "exports" / ".nas-snapshots"
        snapshot = snapshot_folder / f"project-{uuid4().hex}.sqlite"
        create_database_snapshot(path, snapshot)
        assert self._relative_folder is not None
        relative = Path(*self._relative_folder.parts[1:]) / "project.sqlite"
        self._nas_sync.enqueue(snapshot, relative)

    def close(self) -> None:
        with self._lock:
            writer = self._writer
            path = self._path
            self._writer = None
            self._phase_has_records = False
        if writer is not None:
            writer.close()
        if path is not None:
            self._enqueue_nas_snapshot(path)


def numeric_series(
    table: MeasurementTable, names: Iterable[str]
) -> dict[str, tuple[float, ...]]:
    result: dict[str, tuple[float, ...]] = {}
    for name in names:
        values: list[float] = []
        for value in table.column(name):
            try:
                values.append(float(value.replace(",", ".")))
            except ValueError:
                values.append(float("nan"))
        result[name] = tuple(values)
    return result


def wait_for_sync(synchronizer: BackgroundNasSynchronizer, timeout_seconds: float) -> bool:
    """Small test/CLI helper; the UI itself never blocks on NAS."""
    deadline_steps = max(1, int(timeout_seconds / 0.01))
    for _ in range(deadline_steps):
        if synchronizer.pending_count == 0:
            return True
        sleep(0.01)
    return synchronizer.pending_count == 0
