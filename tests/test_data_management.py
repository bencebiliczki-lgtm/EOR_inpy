import csv
import json
from datetime import UTC, datetime
from pathlib import Path

from eor_control.data_management import (
    BackgroundNasSynchronizer,
    MeasurementTable,
    NasSyncQueue,
    ProjectMeasurementWriter,
    export_measurement_csv,
    export_measurement_excel,
    filter_measurement_table_by_stage,
    measurement_stage_segments,
    measurement_stages,
    project_excel_path,
    read_measurement_table,
    read_measurement_tables,
    safe_filename,
)
from eor_control.domain import MeasurementRecord, MeasurementSnapshot, PumpStatus
from eor_control.storage import CsvMeasurementWriter


def record(
    stage: str = "víz", *, recorded_at: datetime | None = None
) -> MeasurementRecord:
    return MeasurementRecord(
        snapshot=MeasurementSnapshot(
            recorded_at=recorded_at or datetime(2026, 7, 13, 12, 30, tzinfo=UTC),
            monotonic_seconds=10.5,
            jacket_pump=PumpStatus(120.5, 1.0, 250.0),
            injection_pump=PumpStatus(100.25, 12.0, 240.0),
            line_pressure_bar=101.75,
            differential_pressure_bar=2.5,
            valve_percent=45.0,
        ),
        injected_volume_ml=10.0,
        active_stage=stage,
    )


def test_project_writer_uses_separate_safe_project_paths(tmp_path: Path) -> None:
    writer = ProjectMeasurementWriter(tmp_path)
    first = writer.select_project(1, 'Első: projekt?', stage_name="víz")
    writer.write(record())
    second = writer.select_project(2, "Második projekt", stage_name="víz")
    writer.write(record())
    writer.close()

    assert first != second
    assert first.name == "Első_projekt_víz_live_raw.csv"
    assert first.parent.parent.name.isdigit()
    assert len(read_measurement_table(first).rows) == 1
    assert len(read_measurement_table(second).rows) == 1
    assert safe_filename('  hibás<>:"/\\|?* név  ') == "hibás_név"


def test_project_writer_creates_portable_json_snapshots(tmp_path: Path) -> None:
    writer = ProjectMeasurementWriter(tmp_path)
    source = writer.select_project_with_metadata(
        7,
        "Kőzet A",
        created_at=datetime(2025, 3, 4, tzinfo=UTC),
        notes="Megjegyzés",
        configuration={"interval": 5},
        calibration_snapshot={"line": [1.0, 5.0, 0.0, 400.0]},
        stages=[{"name": "Hidegvizes", "type": "Hidegvizes"}],
        stage_name="Hidegvizes",
    )
    writer.close()

    assert source.parent.parent.name == "2025"
    project = json.loads((source.parent / "project.json").read_text(encoding="utf-8"))
    assert project["name"] == "Kőzet A"
    assert project["measurement_kind"] == "live"
    assert project["stages"] == [{"name": "Hidegvizes", "type": "Hidegvizes"}]
    assert json.loads(
        (source.parent / "config_snapshot.json").read_text(encoding="utf-8")
    ) == {"interval": 5}


def test_project_folder_uses_hungarian_calendar_date(tmp_path: Path) -> None:
    writer = ProjectMeasurementWriter(tmp_path)

    source = writer.select_project_with_metadata(
        9,
        "Éjfél projekt",
        created_at=datetime(2025, 12, 31, 23, 30, tzinfo=UTC),
    )
    writer.close()

    assert source.parent.parent.name == "2026"
    assert source.parent.name.startswith("2026-01-01_000009_")


def test_user_csv_export_uses_semicolon_and_decimal_comma(tmp_path: Path) -> None:
    writer = ProjectMeasurementWriter(tmp_path)
    source = writer.select_project(1, "Export", stage_name="víz")
    writer.write(record())
    writer.close()
    destination = tmp_path / "export.csv"

    export_measurement_csv(source, destination, decimal_comma=True, delimiter=";")

    with destination.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file, delimiter=";"))
    assert rows[1][2] == "120,5"
    assert rows[1][CsvMeasurementWriter.HEADER.index("line_pressure_bar")] == "101,75"


def test_reader_keeps_legacy_comma_delimited_raw_files_compatible(tmp_path: Path) -> None:
    source = tmp_path / "legacy.csv"
    with source.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "recorded_at_utc",
                "monotonic_seconds",
                "jacket_pressure_bar",
                "jacket_flow_ml_per_hour",
                "jacket_remaining_volume_ml",
                "injection_pressure_bar",
                "injection_flow_ml_per_hour",
                "injection_remaining_volume_ml",
                "injected_volume_ml",
                "line_pressure_bar",
                "differential_pressure_bar",
                "inlet_pressure_bar",
                "valve_percent",
                "active_stage",
                "quality",
                "safety_reasons",
            )
        )
        writer.writerow(["2026-01-01T00:00:00+00:00", *(["1.5"] * 12), "stage", "good", ""])

    table = read_measurement_table(source)

    assert table.rows[0][2] == "1.5"
    assert table.header == CsvMeasurementWriter.HEADER
    assert "inlet_pressure_bar" not in table.header
    assert table.column("jacket_net_volume_ml") == ("",)
    assert table.column("injection_net_volume_ml") == ("1.5",)
    assert table.column("raw_line_pressure_bar") == ("1.5",)
    assert table.column("raw_differential_pressure_bar") == ("1.5",)


def test_measurement_table_filters_stages_and_keeps_repeated_segments() -> None:
    header = CsvMeasurementWriter.HEADER
    stage_index = header.index("active_stage")

    def row(timestamp: str, stage: str) -> tuple[str, ...]:
        values = ["0"] * len(header)
        values[0] = timestamp
        values[stage_index] = stage
        return tuple(values)

    table = MeasurementTable(
        header,
        (
            row("2026-07-13T10:00:00+00:00", "víz"),
            row("2026-07-13T10:01:00+00:00", "víz"),
            row("2026-07-13T10:02:00+00:00", "olaj"),
            row("2026-07-13T10:03:00+00:00", "víz"),
        ),
    )

    assert measurement_stages(table) == ("víz", "olaj")
    assert len(filter_measurement_table_by_stage(table, "víz").rows) == 3
    assert measurement_stage_segments(table) == (
        ("víz", 0, 2),
        ("olaj", 2, 3),
        ("víz", 3, 4),
    )


def test_project_excel_uses_one_charted_worksheet_per_stage(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    writer = ProjectMeasurementWriter(tmp_path)
    water_source = writer.select_project(1, "Excel", stage_name="víz")
    writer.write(record("víz"))
    oil_source = writer.select_project(1, "Excel", stage_name="olaj")
    writer.write(record("olaj"))
    writer.complete_current_phase()
    destination = project_excel_path(water_source, "víz")

    export_measurement_excel(water_source, destination, stage_name="víz")
    export_measurement_excel(oil_source, destination, stage_name="olaj")

    workbook = load_workbook(destination, read_only=False)
    assert destination.name == "Excel.xlsx"
    assert workbook.sheetnames == ["víz", "olaj"]
    assert workbook["víz"]["C2"].value == 120.5
    assert workbook["olaj"]["C2"].value == 120.5
    assert len(workbook["víz"]._charts) == 1
    assert len(workbook["olaj"]._charts) == 1

    writer.select_project(1, "Excel", stage_name="víz")
    writer.write(
        record("víz", recorded_at=datetime(2026, 7, 13, 12, 31, tzinfo=UTC))
    )
    writer.complete_current_phase()
    export_measurement_excel(water_source, destination, stage_name="víz")

    refreshed = load_workbook(destination, read_only=False)
    assert refreshed.sheetnames == ["víz", "olaj"]
    assert refreshed["víz"].max_row == 3
    assert len(refreshed["víz"]._charts) == 1


def test_project_writer_creates_one_raw_csv_per_measurement_stage(
    tmp_path: Path,
) -> None:
    writer = ProjectMeasurementWriter(tmp_path)
    water_path = writer.select_project(1, "Projekt", stage_name="víz")
    writer.write(record("víz"))
    writer.write(record("olaj"))
    writer.close()

    phase_paths = sorted(water_path.parent.glob("*_live_raw.csv"))
    assert [path.name for path in phase_paths] == [
        "Projekt_olaj_live_raw.csv",
        "Projekt_víz_live_raw.csv",
    ]
    assert read_measurement_table(water_path).column("active_stage") == ("víz",)
    assert read_measurement_table(phase_paths[0]).column("active_stage") == ("olaj",)


def test_project_writer_reports_each_completed_measurement_stage_once(
    tmp_path: Path,
) -> None:
    completed: list[tuple[Path, str]] = []
    writer = ProjectMeasurementWriter(
        tmp_path,
        phase_completed=lambda path, stage: completed.append((path, stage)),
    )
    water_path = writer.select_project(1, "Projekt", stage_name="víz")
    writer.write(record("víz"))

    oil_path = writer.select_project(1, "Projekt", stage_name="olaj")

    assert completed == [(water_path, "víz")]
    writer.write(record("olaj"))
    assert writer.complete_current_phase() == oil_path
    assert completed == [(water_path, "víz"), (oil_path, "olaj")]
    assert writer.complete_current_phase() is None
    assert completed == [(water_path, "víz"), (oil_path, "olaj")]


def test_multiple_phase_files_are_combined_only_for_reading(tmp_path: Path) -> None:
    writer = ProjectMeasurementWriter(tmp_path)
    late_path = writer.select_project(1, "Projekt", stage_name="olaj")
    writer.write(
        record("olaj", recorded_at=datetime(2026, 7, 13, 12, 31, tzinfo=UTC))
    )
    early_path = writer.select_project(1, "Projekt", stage_name="víz")
    writer.write(record("víz"))
    writer.close()

    table = read_measurement_tables((late_path, early_path))

    assert table.column("active_stage") == ("víz", "olaj")
    assert sorted(early_path.parent.glob("*_live_raw.csv")) == sorted(
        (early_path, late_path)
    )


def test_disabled_project_writer_never_creates_simulation_files(tmp_path: Path) -> None:
    writer = ProjectMeasurementWriter(tmp_path, enabled=False)

    selected_path = writer.select_project(1, "Szimuláció", stage_name="próba")
    writer.write(record("próba"))
    writer.close()

    assert writer.current_path is None
    assert not selected_path.exists()
    assert not (tmp_path / "projects").exists()


def test_nas_queue_survives_failure_and_resynchronizes(tmp_path: Path) -> None:
    source = tmp_path / "local" / "measurement.csv"
    source.parent.mkdir()
    source.write_text("first", encoding="utf-8")
    queue = NasSyncQueue(tmp_path / "queue.sqlite3")
    synchronizer = BackgroundNasSynchronizer(queue, retry_interval_seconds=60.0)
    blocked_target = tmp_path / "blocked"
    blocked_target.write_text("not a directory", encoding="utf-8")
    synchronizer.configure(enabled=True, target_root=blocked_target)
    synchronizer.enqueue(source, Path("projects") / "measurement.csv")

    assert synchronizer.sync_pending_once() == 0
    assert synchronizer.pending_count == 1
    assert queue.pending()[0].attempts >= 1
    synchronizer.configure(enabled=False, target_root=None)
    synchronizer.close()

    nas_target = tmp_path / "nas"
    reopened_queue = NasSyncQueue(tmp_path / "queue.sqlite3")
    synchronizer = BackgroundNasSynchronizer(reopened_queue, retry_interval_seconds=60.0)
    synchronizer.configure(enabled=True, target_root=nas_target)
    synchronizer.sync_pending_once()
    assert synchronizer.pending_count == 0
    assert (nas_target / "projects" / "measurement.csv").read_text(encoding="utf-8") == "first"
    synchronizer.close()
